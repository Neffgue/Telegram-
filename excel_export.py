"""
Модуль для экспорта данных бота в Excel
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import database

def export_to_excel(filename: str = None):
    """Экспортирует все данные пользователей в Excel файл"""
    if filename is None:
        filename = f"pillow_bot_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Получаем данные из базы
    users_data = database.get_all_users_data()
    interactions = database.get_all_interactions()
    
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    
    # Удаляем дефолтный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Создаем лист с данными пользователей
    ws_users = wb.create_sheet("Пользователи")
    
    # Заголовки для листа пользователей
    headers_users = [
        "ID пользователя",
        "Username",
        "Время напоминания",
        "Часовой пояс",
        "Дата регистрации",
        "Количество принятых таблеток",
        "Дата первого приема",
        "Количество взаимодействий"
    ]
    
    # Задаем стиль для заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Записываем заголовки
    for col_num, header in enumerate(headers_users, 1):
        cell = ws_users.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем данные пользователей
    for row_num, user_data in enumerate(users_data, 2):
        ws_users.cell(row=row_num, column=1, value=user_data[0])  # user_id
        ws_users.cell(row=row_num, column=2, value=user_data[1] or "Не указан")  # username
        ws_users.cell(row=row_num, column=3, value=user_data[2] or "Не установлено")  # reminder_time
        ws_users.cell(row=row_num, column=4, value=user_data[3] or "Не установлен")  # timezone
        ws_users.cell(row=row_num, column=5, value=user_data[4] or "")  # created_at
        ws_users.cell(row=row_num, column=6, value=user_data[5] or 0)  # pills_count
        ws_users.cell(row=row_num, column=7, value=user_data[6] or "")  # first_pill_date
        ws_users.cell(row=row_num, column=8, value=user_data[7] or 0)  # interactions_count
    
    # Автоматически подбираем ширину столбцов для листа пользователей
    for col_num in range(1, len(headers_users) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws_users[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_users.column_dimensions[column_letter].width = adjusted_width
    
    # Создаем лист с историей взаимодействий
    ws_interactions = wb.create_sheet("История взаимодействий")
    
    # Заголовки для листа взаимодействий
    headers_interactions = [
        "ID",
        "ID пользователя",
        "Username",
        "Тип взаимодействия",
        "Данные",
        "Время напоминания",
        "Часовой пояс",
        "Временная метка"
    ]
    
    # Записываем заголовки
    for col_num, header in enumerate(headers_interactions, 1):
        cell = ws_interactions.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем данные взаимодействий
    for row_num, interaction in enumerate(interactions, 2):
        ws_interactions.cell(row=row_num, column=1, value=interaction[0])  # id
        ws_interactions.cell(row=row_num, column=2, value=interaction[1])  # user_id
        ws_interactions.cell(row=row_num, column=3, value=interaction[2] or "Не указан")  # username
        ws_interactions.cell(row=row_num, column=4, value=interaction[3])  # interaction_type
        ws_interactions.cell(row=row_num, column=5, value=interaction[4] or "")  # interaction_data
        ws_interactions.cell(row=row_num, column=6, value=interaction[6] or "")  # reminder_time
        ws_interactions.cell(row=row_num, column=7, value=interaction[7] or "")  # timezone
        ws_interactions.cell(row=row_num, column=8, value=interaction[5])  # timestamp
    
    # Автоматически подбираем ширину столбцов для листа взаимодействий
    for col_num in range(1, len(headers_interactions) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws_interactions[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_interactions.column_dimensions[column_letter].width = adjusted_width
    
    # Создаем лист с историей принятых таблеток
    ws_pills = wb.create_sheet("Принятые таблеточки")
    
    # Получаем данные о принятых таблеточках
    import sqlite3
    conn = sqlite3.connect(database.DATABASE_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT pt.user_id, r.username, pt.date, pt.taken_at, r.reminder_time, r.timezone
        FROM pills_taken pt
        LEFT JOIN reminders r ON pt.user_id = r.user_id
        ORDER BY pt.date DESC, pt.taken_at DESC
    ''')
    pills_data = cursor.fetchall()
    conn.close()
    
    # Заголовки для листа таблеток
    headers_pills = [
        "ID пользователя",
        "Username",
        "Дата",
        "Время приема",
        "Время напоминания",
        "Часовой пояс"
    ]
    
    # Записываем заголовки
    for col_num, header in enumerate(headers_pills, 1):
        cell = ws_pills.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем данные о таблеточках
    for row_num, pill_data in enumerate(pills_data, 2):
        ws_pills.cell(row=row_num, column=1, value=pill_data[0])  # user_id
        ws_pills.cell(row=row_num, column=2, value=pill_data[1] or "Не указан")  # username
        ws_pills.cell(row=row_num, column=3, value=pill_data[2])  # date
        ws_pills.cell(row=row_num, column=4, value=pill_data[3])  # taken_at
        ws_pills.cell(row=row_num, column=5, value=pill_data[4] or "")  # reminder_time
        ws_pills.cell(row=row_num, column=6, value=pill_data[5] or "")  # timezone
    
    # Автоматически подбираем ширину столбцов для листа таблеток
    for col_num in range(1, len(headers_pills) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws_pills[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_pills.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    wb.save(filename)
    return filename

